#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from tkinter import *
from tkinter import filedialog

from vna_scpi import *

import threading
from openpyxl import *


class Measure_thread(threading.Thread):
    def __init__(self, address, start_test= False, test_number = 0,  chart_numbers = 0):
        threading.Thread.__init__(self)

        self.address = address
        self.file_name = ""
        self.time_value = ""
        self.start_test = start_test
        self.test_number = test_number
        self.chart_numbers =  chart_numbers
        self.measure_started = False
        self.data_ready = False

        try:
            self.vna = Vna_measure(self.address)
            self.instrument_info = self.vna.instrument_info()
        except Exception as e:
            print(e)
            self.instrument_info = 'NO CONNECTION \n'
            self.start_test = False

        # opening the existing excel file & create the sheet object
        self.wb = Workbook()
        self.sheet = self.wb.active

    # run all the time
    def run(self):
        try:
            if self.start_test:
                self.measure_started = True
                self.data_ready = False

                #measure vna
                self.measures = []
                for i in range(self.chart_numbers):
                    self.measures.append(self.vna.read_measure(self.test_number, i))

                self.data_ready = True
                self.measure_started = False
        except Exception as e:
            print(e)

    def create_sheet(self):
        xValue = []
        yValue = []

        for i in range(0,len(self.measures),1):
            x, y = self.measures[i]
            xValue.append(x)
            yValue.append(y)

        # - - - - - - - - - - - - - - - - - - - - -
        # Create sheet
        self.sheet.cell(row=1, column=1).value = self.file_name
        self.sheet.cell(row=1, column=2).value = self.time_value

        for i in range(0, len(xValue), 1):
            self.sheet.cell(row=2, column= (i * 2) + 1).value = 'data: ' + str(i + 1)

            self.sheet.cell(row=3, column= (i * 2) + 1).value = 'x'
            self.sheet.cell(row=3, column= (i * 2) + 2).value = 'y'

            for j in range(0, len(xValue[0]), 1):
                self.sheet.cell(row=j + 4, column= (i * 2) + 1).value = xValue[i][j]
                self.sheet.cell(row=j + 4, column= (i * 2) + 2).value = yValue[i][j]

        file_position = filedialog.asksaveasfilename(initialdir = "/",initialfile=self.file_name, title = "Select file",filetypes = (("Microsoft Excel Worksheet","*.xlsx"),("all files","*.*")), defaultextension = ' ')
        self.wb.save(file_position)
