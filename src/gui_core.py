# -*- coding: utf-8 -*-

import os

import settings
from user_gui import *
from vna_scpi import *

from openpyxl import *

import time
from time import gmtime, strftime

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.qt_compat import QtCore, QtWidgets, is_pyqt5
from matplotlib.backends.backend_qt5agg import (FigureCanvas, NavigationToolbar2QT as NavigationToolbar)
from matplotlib.figure import Figure


#==============================================================================#
def file_open(self):
    name = QtWidgets.QFileDialog.getOpenFileName(MainWindow, 'Open File')  # Returns a tuple
    #print(name)  # For debugging
    if len(name[0]) > 0:
        with open(name[0], 'r') as file:
            text = file.read()
            self.textEdit.setText(text)

def file_save(self):
    title = self.serialName.text() + self.serialNumber.text() + self.addDetails.text()
    name = QtWidgets.QFileDialog.getSaveFileName(MainWindow, 'Save file', os.path.join(str(os.getenv('HOME')), title), 'Microsoft Excel Worksheet(*.xlsx);; Text file(*.txt);; All files(*.*)')  # Returns a tuple
    #print(name)

    if self.tabWidget.currentIndex() == 3:
        if len(name[0]) > 0:
            with open(name[0], 'w') as file:
                #text = self.textEdit.toPlainText()  # Plain text without formatting
                text = self.textEdit.toHtml()  # Rich text with formatting (font, color, etc.)
                file.write(text)
    else:
        try:
            name, _ = name

            xValue = []
            yValue = []

            measures = self.vna_measure.measures

            self.wb = Workbook()
            self.sheet = self.wb.active

            for i in range(len(measures)):
                x, y = measures[i]
                xValue.append(x)
                yValue.append(y)

            # Create sheet
            self.sheet.cell(row=1, column=1).value = "test"
            self.sheet.cell(row=1, column=2).value = "test 2"

            for i in range(0, len(xValue), 1):
                self.sheet.cell(row=2, column= (i * 2) + 1).value = 'data: ' + str(i + 1)
                self.sheet.cell(row=3, column= (i * 2) + 1).value = 'x'
                self.sheet.cell(row=3, column= (i * 2) + 2).value = 'y'
                for j in range(0, len(xValue[0]), 1):
                    self.sheet.cell(row=j + 4, column= (i * 2) + 1).value = xValue[i][j]
                    self.sheet.cell(row=j + 4, column= (i * 2) + 2).value = yValue[i][j]

                self.wb.save(name)
        except Exception as e:
            print(e)

def file_quit(self):
    decision = QtWidgets.QMessageBox.question(MainWindow, 'Question',
                   'Sure to quit?',
                   QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
    if decision == QtWidgets.QMessageBox.Yes:
        sys.exit()

def edit_font(self):  # Applies on all text
    font = self.textEdit.font()  # Get the currently used font
    newfont, valid = QtWidgets.QFontDialog.getFont(font)
    if valid:
        self.textEdit.setFont(newfont)

def edit_color(self):  # Applies on selected text only
    color = self.textEdit.textColor()
    newcolor = QtWidgets.QColorDialog.getColor(color)
    if newcolor.isValid():
        self.textEdit.setTextColor(newcolor)

def enlarge_window(self):
    if self.checkBox.isChecked():
        MainWindow.setGeometry(100, 100, 600, 400)
        #MainWindow.setWindowTitle("PyQT tuts!")
        """
        import os
        scriptDir = os.path.dirname(os.path.realpath(__file__))
        MainWindow.setWindowIcon(QtGui.QIcon(scriptDir + os.path.sep + 'images/icon.ico'))
        """
        """
        from PyQt5.QtWidgets import QStyleFactory
        app.setStyle(QStyleFactory.create('Fusion'))
        """
    else:
        MainWindow.setGeometry(100, 100, 400, 300)


#==============================================================================#
def check_input(self):
    self.actionOpen.triggered.connect(self.file_open)
    self.actionSave.triggered.connect(self.file_save)
    self.actionQuit.triggered.connect(self.file_quit)
    self.actionFont.triggered.connect(self.edit_font)
    self.actionColor.triggered.connect(self.edit_color)

    self.connectButton.clicked.connect(self.connect_instrument)
    self.startMeasure.clicked.connect(self.start_measure)

    self.checkBox.stateChanged.connect(self.enlarge_window)
    self.saveReference.stateChanged.connect(self.save_reference)
    self.addTrace.clicked.connect(self.add_trace)
    self.removeTrace.clicked.connect(self.remove_trace)
    self.saveSparameters.clicked.connect(self.save_s_parameters)

def connect_instrument(self, current_index = 0):
    address = self.instrumentAddress.text()

    self.vna_measure = Vna_measure(address, current_index)

    self.instrument_timer = QtCore.QTimer()
    self.instrument_timer.timeout.connect(self.instrument_refresh)
    self.instrument_timer.start(1000)

    self.progressBar.setValue(0)

def start_measure(self):
    self.connect_instrument(self.tabWidget.currentIndex())
    #print(self.tabWidget.currentIndex())
    counter = self.lcdNumber.value() + 1
    self.lcdNumber.display(counter)

def save_reference(self):
    self.update_plot()

def add_trace(self):
    self.saveRef = True
    self.update_plot()

def remove_trace(self):
    self.delRef = True
    self.update_plot()

def save_s_parameters(self):
    print("S-parameters")
    try:
        title = self.serialName.text() + self.serialNumber.text() + self.addDetails.text() + "_s-param"
        name = QtWidgets.QFileDialog.getSaveFileName(MainWindow, 'Save file', os.path.join(str(os.getenv('HOME')), title), 'All files(*.*)')  # Returns a tuple
        name, _ = name

        file = open(name,"a")
        file.write(self.vna_measure.s_parameters)
        file.close()
        print('File saved')

        # export csv file
        title = self.serialName.text() + self.serialNumber.text() + self.addDetails.text() + "all_traces"
        name = QtWidgets.QFileDialog.getSaveFileName(MainWindow, 'Save file', os.path.join(str(os.getenv('HOME')), title), 'All files(*.*)')  # Returns a tuple
        name, _ = name

        file = open(name,"a")
        file.write(self.vna_measure.all_traces)
        file.close()
        print('File saved')

    except Exception as e:
        print(e)

#==============================================================================#
def instrument_refresh(self):
    try:
        self.remoteConnectionLabel.setText(self.vna_measure.instrument_info)
        #self.update_plot()

        bar_value = self.progressBar.value()
        if bar_value < 100:
            bar_value += 5
            self.progressBar.setValue(bar_value)

        if self.vna_measure.data_ready == True:
            self.vna_measure.data_ready = False

            self.update_plot()
            self.progressBar.setValue(100)

            if self.autoSave.isChecked():
                self.file_save()

    except Exception as e:
        print(e)

def update_time(self):
    self.timeLabel.setText(strftime("%d %b %Y %H:%M:%S", gmtime()))
    self.timer = QtCore.QTimer()
    self.timer.timeout.connect(self.update_time)
    self.timer.start(1000)


#==============================================================================#
def create_canvas(self):
    """
    static_canvas = FigureCanvas(Figure(figsize=(5, 3)))
    self.plotTest.addWidget(static_canvas)
    MainWindow.addToolBar(NavigationToolbar(static_canvas, MainWindow))

    self._static_ax = static_canvas.figure.subplots()
    t = np.linspace(0, 10, 501)
    self._static_ax.plot(t, np.tan(t), ".")
    """

    dynamic_canvas = FigureCanvas(Figure(figsize=(5, 3)))
    self.plotTest.addWidget(dynamic_canvas)
    self.plotTest.addWidget(NavigationToolbar(dynamic_canvas, MainWindow))

    self._dynamic_ax = dynamic_canvas.figure.subplots()

    self._timer = dynamic_canvas.new_timer(
        100, [(self.update_canvas, (), {})])
    self._timer.start()

def update_canvas(self):
    self._dynamic_ax.clear()
    t = np.linspace(0, 10, 101)
    # Shift the sinusoid as a function of time.
    self._dynamic_ax.plot(t, np.sin(t + time.time()))
    self._dynamic_ax.figure.canvas.draw()


#==============================================================================#
def create_plot(self):
    plt.style.use('seaborn-whitegrid')

    number_of_plots = len(settings.plot_names)

    self.plot = [[] for i in range(number_of_plots)]
    self.fig = [[] for i in range(number_of_plots)]
    self.toolbar = [[] for i in range(number_of_plots)]

    # initialize fig
    for i in range(number_of_plots):
        self.fig[i] =Figure(figsize=(12,7))

    # Figures
    for index in range(number_of_plots):
        subplot_number = len(settings.plot_names[index])
        for i in range(subplot_number):
            # auto adapt plot number
            subplot_columns = (subplot_number // 3) + (subplot_number % 3)
            subplot_rows = subplot_number // 2
            subplot = self.fig[index].add_subplot(subplot_rows, subplot_columns, i + 1)
            self.plot[index].append(subplot)

    # auto adj
    for i in range(number_of_plots):
        self.fig[i].tight_layout()

    # Canvas & toolbar
    thisFigure = FigureCanvas(self.fig[0])
    self.plotTest_2.addWidget(thisFigure)
    self.plotTest_2.addWidget(NavigationToolbar(thisFigure, MainWindow))

    thisFigure = FigureCanvas(self.fig[1])
    self.plotTest_3.addWidget(thisFigure)
    self.plotTest_3.addWidget(NavigationToolbar(thisFigure, MainWindow))

def update_plot(self):
    try:
        #code to review
        if self.plotRef != self.tabWidget.currentIndex():
            self.plotRef = -1
            self.delRef = True
            self.saveReference.setChecked(False)

        self.plotRef = self.tabWidget.currentIndex()

        if self.saveReference.isChecked():
            self.saveRef = True

        # return wich test you have selected
        selected_frame_number = self.vna_measure.test_type - 1

        if (len(self.vna_measure.measures)) < (len(settings.plot_names[selected_frame_number])):
            channel_number = len(self.vna_measure.measures)
        else:
            channel_number = len(settings.plot_names[selected_frame_number])

        xValue = []
        yValue = []

        for i in range(channel_number):
            x, y = self.vna_measure.measures[i]
            xValue.append(x)
            yValue.append(y)
            # clean plot line
            self.plot[selected_frame_number][i].clear()
            # set data on plot
            self.plot[selected_frame_number][i].plot(xValue[i], yValue[i])

        if self.saveRef == True:
            self.xRef.append(xValue)
            self.yRef.append(yValue)
            self.saveRef = False

        if self.delRef == True:
            self.xRef = []
            self.yRef = []
            self.delRef = False

        for j in range(len(self.xRef)):
            for i in range(channel_number):
                self.plot[selected_frame_number][i].plot(self.xRef[j][i], self.yRef[j][i])

        # Set names on plot
        for i in range(len(settings.plot_names[selected_frame_number])):
            self.plot[selected_frame_number][i].set_title(settings.plot_names[selected_frame_number][i][0])
            self.plot[selected_frame_number][i].set_xlabel(settings.plot_names[selected_frame_number][i][1])
            self.plot[selected_frame_number][i].set_ylabel(settings.plot_names[selected_frame_number][i][2])
            #self.plot[selected_frame_number][i].grid()

        # autoadapt
        self.fig[selected_frame_number].tight_layout()
        # update plot
        self.fig[selected_frame_number].canvas.draw()

        # markers
        #Single_marker(self.others_frame, self.canvas[selected_frame_number], self.plot[selected_frame_number], xValue, yValue, len(settings.plot_names[selected_frame_number]), self.plot_markers)

    except Exception as e:
        print(e)


#==============================================================================#
Ui_MainWindow.check_input = check_input
Ui_MainWindow.file_open = file_open
Ui_MainWindow.file_save = file_save
Ui_MainWindow.file_quit = file_quit
Ui_MainWindow.edit_font = edit_font
Ui_MainWindow.edit_color = edit_color
Ui_MainWindow.enlarge_window = enlarge_window

Ui_MainWindow.connect_instrument = connect_instrument
Ui_MainWindow.start_measure = start_measure

Ui_MainWindow.save_reference = save_reference
Ui_MainWindow.add_trace = add_trace
Ui_MainWindow.remove_trace = remove_trace

Ui_MainWindow.save_s_parameters = save_s_parameters

Ui_MainWindow.instrument_refresh = instrument_refresh
Ui_MainWindow.update_time = update_time

Ui_MainWindow.create_canvas = create_canvas
Ui_MainWindow.update_canvas = update_canvas

Ui_MainWindow.create_plot = create_plot
Ui_MainWindow.update_plot = update_plot


Ui_MainWindow.xRef = []
Ui_MainWindow.yRef = []
Ui_MainWindow.saveRef = False
Ui_MainWindow.delRef = False
Ui_MainWindow.plotRef = -1


#==============================================================================#
#if __name__ == "__main__":
import sys
app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
ui.update_time()
ui.check_input()
ui.create_canvas()
ui.create_plot()
MainWindow.show()
sys.exit(app.exec_())
