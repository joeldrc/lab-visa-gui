import threading

from visa_scpi import Vna_measure

import time
from time import gmtime, strftime

from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog

from openpyxl import *

from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.figure import Figure
import matplotlib.pyplot as plt


class Progress():
    # threaded progress bar for tkinter gui
    def __init__(self, parent, row, columnspan, sticky, padx, pady):
        self.maximum = 100
        self.interval = 10
        self.progressbar = ttk.Progressbar(parent, orient=tk.HORIZONTAL, mode="indeterminate", maximum=self.maximum)
        self.progressbar.grid(row= row, columnspan=columnspan, sticky=sticky, padx=padx, pady=pady)
        self.thread = threading.Thread()
        self.thread.__init__(target=self.progressbar.start(self.interval), args=())
        self.pb_clear()


    def pb_stop(self):
        """ stops the progress bar """
        if not self.thread.isAlive():
            VALUE = self.progressbar["value"]
            self.progressbar.stop()
            self.progressbar["value"] = VALUE


    def pb_start(self):
        """ starts the progress bar """
        if not self.thread.isAlive():
            VALUE = self.progressbar["value"]
            self.progressbar.configure(mode="indeterminate", maximum=self.maximum, value=VALUE)
            self.progressbar.start(self.interval)


    def pb_clear(self):
        """ stops the progress bar """
        if not self.thread.isAlive():
            self.progressbar.stop()
            self.progressbar.configure(mode="determinate", value=0)


    def pb_complete(self):
        """ stops the progress bar and fills it """
        if not self.thread.isAlive():
            self.progressbar.stop()
            self.progressbar.configure(mode="determinate", maximum=self.maximum, value=self.maximum)


class My_thread(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

        self.file_name = "_null_"
        self.data_ready = False
        self.save_data = True

        self.vna = Vna_measure('TCPIP::CFO-MD-BQPVNA1::INSTR')
        self.instrument_info = self.vna.instrument_info()

        # opening the existing excel file & create the sheet object
        self.wb = Workbook()
        self.sheet = self.wb.active


    # run all the time
    def run(self):
        self.data_ready = False
        
        # masure vna
        self.measure0 = self.vna.read_measure(0)
        self.measure1 = self.vna.read_measure(1)
        self.measure2 = self.vna.read_measure(2)
        self.measure3 = self.vna.read_measure(3)
        self.measure4 = self.vna.read_measure(4)

        self.data_ready = True                

        if self.save_data:
            self.create_sheet()


    def create_sheet(self):
        # masure vna
        xValue0, yValue0 = self.measure0
        xValue1, yValue1 = self.measure1
        xValue2, yValue2 = self.measure2
        xValue3, yValue3 = self.measure3
        xValue4, yValue4 = self.measure4

        # - - - - - - - - - - - - - - - - - - - - -
        # Create sheet
        """
        # resize the width of columns in excel spreadsheet
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 30
        """
        
        self.sheet.cell(row=1, column=1).value = self.file_name

        self.sheet.cell(row=1, column=2).value =strftime("%d %b %Y %H:%M:%S", gmtime())

        self.sheet.cell(row=2, column=1).value = 'x'
        self.sheet.cell(row=2, column=2).value = 'y'

        self.sheet.cell(row=2, column=4).value = 'x'
        self.sheet.cell(row=2, column=5).value = 'y'

        self.sheet.cell(row=2, column=7).value = 'x'
        self.sheet.cell(row=2, column=8).value = 'y'

        self.sheet.cell(row=2, column=10).value = 'x'
        self.sheet.cell(row=2, column=11).value = 'y'

        self.sheet.cell(row=2, column=13).value = 'x'
        self.sheet.cell(row=2, column=14).value = 'y'

        for i in range(0, len(xValue0), 1):
            self.sheet.cell(row=i + 3, column=1).value = xValue0[i]
            self.sheet.cell(row=i + 3, column=2).value = yValue0[i]

            self.sheet.cell(row=i + 3, column=4).value = xValue1[i]
            self.sheet.cell(row=i + 3, column=5).value = yValue1[i]

            self.sheet.cell(row=i + 3, column=7).value = xValue2[i]
            self.sheet.cell(row=i + 3, column=8).value = yValue2[i]

            self.sheet.cell(row=i + 3, column=10).value = xValue3[i]
            self.sheet.cell(row=i + 3, column=11).value = yValue3[i]

            self.sheet.cell(row=i + 3, column=13).value = xValue4[i]
            self.sheet.cell(row=i + 3, column=14).value = yValue4[i]
                    
        file_position = filedialog.asksaveasfilename(initialdir = "/",initialfile=self.file_name, title = "Select file",filetypes = (("Microsoft Excel Worksheet","*.xlsx"),("all files","*.*")), defaultextension = ' ')
        self.wb.save(file_position)

        
class User_gui(tk.Frame):
    
    def __init__(self, parent,):
        self.window = parent
        #self.window.geometry('600x600')
        #self.window.configure(background='gray')
        
        # variables
        self.plot_reference = False
        self.plot_saveRef = False

        self.myThreadOb1 = My_thread()
        #self.myThreadOb1.start()
        #myThreadOb1.join()
             
        self.window.title("TEST GUI - V.1.0")
        self.create_widgets()
        
               
    def focus(self, event):
        self.start_test()


    # - - - - - - - - - - - - - - - - - - - - -
    # menuBar	
    def show_info(self):
        messagebox.showinfo(title = 'About Me!', message = 'joel.daricou@cern.ch 2018')


    def save_file(self):    
        #file_name = self.serial_name_field.get() + '_' + self.serial_number_field.get() + '_' + self.details_field.get()
        return


    def open_file(self):
        #self.file_position = filedialog.askopenfile().name
        return
        

    def close_file(self):
        exit = messagebox.askyesno(title = 'Quit?', message = 'Are you sure?')
        if exit > 0:
            self.window.destroy()


    # - - - - - - - - - - - - - - - - - - - - -
    # autoupdate
    def update_screen(self):
        if self.myThreadOb1.data_ready:
            self.prog_bar.pb_complete()
            self.create_plot()
        
        self.clock.config(text=strftime("%d %b %Y %H:%M:%S", gmtime()))        
        self.clock.after(1000, self.update_screen)


    # - - - - - - - - - - - - - - - - - - - - -
    # user pannel   
    def panel_led(self, color):
        self.circle_canvas.create_oval(10, 10, 40, 40, width=0, fill=color)


    def save_ref(self):                      
        # invert the value
        self.plot_reference = not self.plot_reference
        if self.plot_reference == True:
            self.save_ref.config(text='Remove Ref.')
            self.plot_saveRef = True
        else:
            self.save_ref.config(text='Set Ref.')


    def start_test(self):
        self.prog_bar.pb_start()
        
        self.myThreadOb1 = My_thread()
        self.myThreadOb1.start()
        
        self.save_data()
        self.myThreadOb1.val = True     
        self.myThreadOb1.file_name = self.serial_name_field.get() + '_' + self.serial_number_field.get() + '_' + self.details_field.get()
        #print(self.myThreadOb1.isAlive())


    def save_data(self):
        self.myThreadOb1.save_data = self.var.get()
        
        
    def set_axis_name(self):
        # set axis names
        self.plot0.set_title('S21 - delay')
        self.plot0.set_xlabel('freq')
        self.plot0.set_ylabel('dB')

        self.plot1.set_title('S21 - dB')
        self.plot1.set_xlabel('freq')
        self.plot1.set_ylabel('dB')

        self.plot2.set_title('S11 - SWR')
        self.plot2.set_xlabel('freq')
        self.plot2.set_ylabel('dB')

        self.plot3.set_title('S22 - SWR')
        self.plot3.set_xlabel('freq')
        self.plot3.set_ylabel('dB')

        self.plot4.set_title('S11 - TDR')
        self.plot4.set_xlabel('delay')
        self.plot4.set_ylabel('dB')


    def create_plot(self):
        # masure vna
        xValue0, yValue0 = self.myThreadOb1.measure0
        xValue1, yValue1 = self.myThreadOb1.measure1
        xValue2, yValue2 = self.myThreadOb1.measure2
        xValue3, yValue3 = self.myThreadOb1.measure3
        xValue4, yValue4 = self.myThreadOb1.measure4

        if self.plot_saveRef == True:
            # set reference data on plot
            self.xRef0, self.yRef0 = xValue0, yValue0
            self.xRef1, self.yRef1 = xValue1, yValue1
            self.xRef2, self.yRef2 = xValue2, yValue2
            self.xRef3, self.yRef3 = xValue3, yValue3
            self.xRef4, self.yRef4 = xValue4, yValue4
            saveRef = False           
          
        # clean plot line
        self.plot0.cla()
        self.plot1.cla()
        self.plot2.cla()
        self.plot3.cla()
        self.plot4.cla()

        # set axis names
        self.set_axis_name()
      
        # set data on plot
        self.plot0.plot(xValue0, yValue0)        
        self.plot1.plot(xValue1, yValue1)
        self.plot2.plot(xValue2, yValue2)
        self.plot3.plot(xValue3, yValue3)
        self.plot4.plot(xValue4, yValue4)

        if self.plot_reference == True:     
            self.plot0.plot(self.xRef0, self.yRef0)        
            self.plot1.plot(self.xRef1, self.yRef1)
            self.plot2.plot(self.xRef2, self.yRef2)
            self.plot3.plot(self.xRef3, self.yRef3)
            self.plot4.plot(self.xRef4, self.yRef4)

        # update plot    
        self.canvas.draw()

           
    def create_widgets(self):       
        # - - - - - - - - - - - - - - - - - - - - -
        # menuBar
        myMenuBar = Menu (self.window)

        myFileMenu = Menu (myMenuBar , tearoff = 0)
        myFileMenu.add_command(label = "Exit", command = self.close_file)
        myFileMenu.add_command(label = "Open", command = self.open_file)
        myFileMenu.add_command(label = "Save as", command = self.save_file)
        myMenuBar.add_cascade(label = "File", menu = myFileMenu)

        myFileMenu = Menu (myMenuBar , tearoff = 0)
        myFileMenu.add_command(label = "Info", command = self.show_info)
        myMenuBar.add_cascade(label = "Help", menu = myFileMenu)

        self.window.config(menu = myMenuBar)

        # create some room around all the internal frames
        self.window['padx'] = 10
        self.window['pady'] = 10      

        # - - - - - - - - - - - - - - - - - - - - -
        # title
        instrument_name, instrument_address = self.myThreadOb1.instrument_info
        str_instrument_info = instrument_name + "\n" + instrument_address
        
        labeled_frame_label = ttk.Label(self.window, text=str_instrument_info)
        labeled_frame_label.grid(row=0, column=0, sticky=W, padx=10, pady=5)

        # - - - - - - - - - - - - - - - - - - - - -    
        # user data frame
        frame = ttk.LabelFrame(self.window, text="USER DATA", relief=tk.RIDGE)
        frame.grid(row=1, column=1, sticky = tk.E + tk.W + tk.N + tk.S, padx=10, pady=10)

        # display label
        name = Label(frame, text='Name (User):')
        name.grid(row=1, column=0, sticky = W)
        self.name_field = Entry(frame)
        self.name_field.grid(row=1, column=1, padx=5, pady = 5)

        serial_name = Label(frame, text='Serial name:')
        serial_name.grid(row=2, column=0, sticky = W)
        self.serial_name_field = Entry(frame)
        self.serial_name_field.grid(row=2, column=1, padx=5, pady = 5)

        serial_number = Label(frame, text='Serial num.:')
        serial_number.grid(row=3, column=0, sticky = W)
        self.serial_number_field = Entry(frame)
        self.serial_number_field.grid(row=3, column=1, padx=5, pady = 5)

        details = Label(frame, text='Add details:')
        details.grid(row=4, column=0, sticky = W)
        self.details_field = Entry(frame)
        self.details_field.grid(row=5, column=0, sticky = E + W, columnspan=2, padx=5, pady = 5)

        # display check button
        self.var = IntVar(value=1)
        self.check_save_file = Checkbutton(frame, text = "Save data after measure", variable=self.var, command= self.save_data)
        self.check_save_file.grid(row=10, column=0, padx=0, pady=5)
        
        # display button
        self.save_ref = Button(frame, text='Save ref.', fg='Black', command= self.save_ref)
        self.save_ref.grid(row=11, column=0, columnspan=2, sticky = W, padx=20, pady = 5)
        
        submit = Button(frame, text='START MEASURE', fg='Black', command= self.start_test)
        submit.grid(row=11, column=1, columnspan=2, sticky = E, padx=20, pady = 5)

        # create loading bar
        self.prog_bar = Progress(frame, row=30, columnspan=2, sticky = tk.E + tk.W + tk.N + tk.S, padx=5, pady=10)

        # display time
        self.clock = Label(frame)
        self.clock.grid(row=31, column=0, sticky = tk.W + tk.N + tk.S, padx=5, pady=5)

        # - - - - - - - - - - - - - - - - - - - - -
        # plot setup
        plt.style.use('bmh')
        
        self.fig = Figure()       
        self.fig = plt.gcf()
        DPI = self.fig.get_dpi()
        self.fig.set_size_inches(1280.0/float(DPI), 720.0/float(DPI))
        
        # subplot
        self.plot0 = self.fig.add_subplot(231)     
        self.plot1 = self.fig.add_subplot(232)    
        self.plot2 = self.fig.add_subplot(233)      
        self.plot3 = self.fig.add_subplot(234)     
        self.plot4 = self.fig.add_subplot(235)

        # set axis names
        self.set_axis_name()
        
        # autoadapt
        plt.tight_layout()

        # draw        
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.window)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(row=1, column=0, sticky=tk.E + tk.W + tk.N + tk.S, padx=10, pady=10)

        # - - - - - - - - - - - - - - - - - - - - -
        # event
        self.update_screen()
               
        # whenever the enter key is pressed then call the focus1 function
        self.window.bind('<Return>', self.focus)

        
